import tkinter as tk
from tkinter import filedialog
import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(title="Selecione a planilha do Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
print(f"Caminho do arquivo selecionado: {file_path}")

if file_path:
    try:
        # Lendo a planilha
        df = pd.read_excel(file_path)

        print("Primeiras linhas do DataFrame original:")
        print(df.head())

        print("Nomes das colunas no DataFrame:", df.columns)

        titulo_column = 'Título'
        data_column = 'Data do Evento'
        descricao_column = 'Descrição do Evento'
        autoridade_column = 'Autoridade'

        # Verificando colunas
        if titulo_column not in df.columns or data_column not in df.columns or autoridade_column not in df.columns or descricao_column not in df.columns:
            raise ValueError(f"Colunas '{titulo_column}', '{data_column}', '{descricao_column}' ou '{autoridade_column}' não encontradas no DataFrame.")

        def extrair_local(descricao):
            padrao_local = r'(São Paulo|Brasília|Nova Iorque|Índia|Washington|Londres)'  
            match = re.search(padrao_local, descricao, re.IGNORECASE)
            if match:
                return match.group(1).strip()
            else:
                return None
        df['Local'] = df[descricao_column].apply(lambda x: extrair_local(str(x)))
        df[data_column] = pd.to_datetime(df[data_column], errors='coerce', format='%d/%m/%Y').dt.date
        print("Primeiras linhas com a nova coluna 'Local' e data ajustada:")
        print(df.head())

        start_date = pd.to_datetime('2023-02-28')
        end_date = pd.to_datetime('2024-06-28')

        filtered_df = df[(df[autoridade_column] == '01 - Roberto Campos Neto') & 
                         (df[data_column] >= start_date) & 
                         (df[data_column] <= end_date) &
                         (~df['Local'].isnull())]
        if filtered_df.empty:
            print("Nenhum dado encontrado para o filtro especificado.")
        else:
            filtered_df.drop(columns=[titulo_column], inplace=True)

            # resultado filtrado
            filtered_df.to_excel('agenda_filtrada.xlsx', index=False)
            print("Arquivo filtrado salvo como 'agenda_filtrada.xlsx'")

            # Estilizando
            wb = openpyxl.load_workbook('agenda_filtrada.xlsx')
            ws = wb.active
            ws.insert_rows(1)
            ws['A1'] = 'AGENDA DO PRESIDENTE DO BACEN'
            ws['A1'].font = Font(size=12, bold=True, color="FFFFFF")
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A1'].fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(filtered_df.columns))

            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col in ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=len(filtered_df.columns)):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

            for col in ws.iter_cols(min_row=3, min_col=1, max_col=len(filtered_df.columns)):
                for cell in col:
                    if cell.column != 2:
                        cell.alignment = Alignment(horizontal='center')

            for col in ws.iter_cols(min_col=1, max_col=len(filtered_df.columns)):
                max_length = 40
                for cell in col:
                    if not isinstance(cell, openpyxl.cell.MergedCell):
                        ws.column_dimensions[cell.column_letter].width = max_length

            # Salvando arquivo estilizado
            wb.save('agenda_filtrada_stylized.xlsx')
            print("Arquivo estilizado salvo como 'agenda_filtrada_stylized.xlsx'")
    
    except FileNotFoundError:
        print("Arquivo não encontrado. Verifique o caminho e o nome do arquivo.")
    except ValueError as ve:
        print(ve)
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
else:
    print("Nenhum arquivo foi selecionado.")